import openpyxl
import json

wb = openpyxl.load_workbook('c:/Users/docte/Downloads/ASF_ BP Canevas (1).xlsx', data_only=True)

results = {}

# B.1. P&L
ws = wb['B.1. P&L']
pnl = {}
for r in range(1, 40):
    label = ws.cell(row=r, column=2).value
    if label:
        vals = []
        for c in range(3, 10):
            v = ws.cell(row=r, column=c).value
            vals.append(str(v) if v is not None else '')
        pnl[str(label)] = vals
results['pnl'] = pnl

# A.2. Chiffre d'Affaires
ws = wb["A.2. Chiffre d'Affaires "]
ca = {}
for r in range(1, 120):
    label = ws.cell(row=r, column=2).value
    if label:
        vals = []
        for c in range(3, 15):
            v = ws.cell(row=r, column=c).value
            vals.append(str(v) if v is not None else '')
        ca[f'Row{r}_{str(label)}'] = vals
results['ca'] = ca

# A.4. Masse Salariale
ws = wb['A.4. Masse Salariale']
sal = {}
for r in range(1, 80):
    label = ws.cell(row=r, column=1).value
    label2 = ws.cell(row=r, column=2).value
    if label or label2:
        vals = []
        for c in range(1, 12):
            v = ws.cell(row=r, column=c).value
            vals.append(str(v) if v is not None else '')
        sal[f'Row{r}'] = vals
results['salary'] = sal

# C. Synthèse Financement
ws = wb['C. Synthèse Financement ']
synth = {}
for r in range(1, 60):
    label = ws.cell(row=r, column=2).value
    if label:
        vals = []
        for c in range(1, 12):
            v = ws.cell(row=r, column=c).value
            vals.append(str(v) if v is not None else '')
        synth[f'Row{r}_{str(label)}'] = vals
results['synth'] = synth

with open('c:/Users/docte/Downloads/unimove-dz-saa-s-2/excel_data.json', 'w', encoding='utf-8') as f:
    json.dump(results, f, ensure_ascii=False, indent=2)

print('Done')
