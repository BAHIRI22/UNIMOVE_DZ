import openpyxl
import json

wb = openpyxl.load_workbook('c:/Users/docte/Downloads/ASF_ BP Canevas (1).xlsx', data_only=True)

results = {}

# Read A.2. CA completely
ws = wb["A.2. Chiffre d'Affaires "]
ca_rows = []
for r in range(1, ws.max_row + 1):
    vals = []
    has_data = False
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=r, column=c).value
        s = str(v).strip() if v is not None else ''
        vals.append(s)
        if s and s != 'None':
            has_data = True
    if has_data:
        ca_rows.append({'row': r, 'values': vals})
results['ca_full'] = ca_rows

# Read A.4 completely
ws = wb['A.4. Masse Salariale']
sal_rows = []
for r in range(1, ws.max_row + 1):
    vals = []
    has_data = False
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=r, column=c).value
        s = str(v).strip() if v is not None else ''
        vals.append(s)
        if s and s != 'None':
            has_data = True
    if has_data:
        sal_rows.append({'row': r, 'values': vals})
results['salary_full'] = sal_rows

# Search for market size / TAM / SAM / SOM / partenariat / université in all sheets
found_terms = []
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    for r in range(1, min(ws.max_row + 1, 300)):
        for c in range(1, min(ws.max_column + 1, 20)):
            v = ws.cell(row=r, column=c).value
            if v and isinstance(v, str):
                lower_v = v.lower()
                if any(term in lower_v for term in ['tam', 'sam', 'som', 'partenariat', 'université', 'étudiant', 'universite', 'effectif', 'headcount', 'etp', 'nbre salarié', 'nombre de salarié', 'transport', 'étudiants', 'abonnement']):
                    row_vals = []
                    for cc in range(1, min(ws.max_column + 1, 15)):
                        vv = ws.cell(row=r, column=cc).value
                        row_vals.append(str(vv).strip() if vv is not None else '')
                    found_terms.append({'sheet': sheet_name, 'row': r, 'col': c, 'text': v, 'row_vals': row_vals})
                    break
results['found_terms'] = found_terms

with open('c:/Users/docte/Downloads/unimove-dz-saa-s-2/excel_data2.json', 'w', encoding='utf-8') as f:
    json.dump(results, f, ensure_ascii=False, indent=2)

print('Done')
